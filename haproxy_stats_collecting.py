import requests  # pip3 install requests
import time
import sys
from openpyxl import Workbook  # pip3 install openpyxl
from openpyxl import load_workbook
from datetime import datetime, timedelta


def fetch_data(url):
    """
    Получение данных
    :param url: URL строки
    :return: текст запроса при успехе и ничего при ошибке
    """
    try:
        response = requests.get(url)
        if response.status_code == 200:
            return response.text
        else:
            print(f"Ошибка: Не удалось получить данные. Status code: {response.status_code}")
    except requests.exceptions.RequestException as error:
        print(f"Ошибка: {error}")
    return None


servers_strings = [   # список доп серверов, которые мониторить
    "back::Thin_Client_Users:88,nn-sed-web1.nnov.ru",
    "back::Thin_Client_Users:88,nn-sed-web2.nnov.ru",
    "back::Thin_Client_Users:88,nn-sed-web3.nnov.ru",
    "back::Thin_Client_Users:88,nn-sed-web4.nnov.ru",
    "back::Thin_Client_Users:88,nn-sed-web5.nnov.ru",
    "back::Thin_Client_Users:88,nn-sed-web7.nnov.ru",
    "back::Thin_Client_Users:88,nn-sed-web8.nnov.ru",
    "back::Thin_Client_Users:88,nn-sed-web9.nnov.ru",
    "back::Thin_Client_Users:88,nn-sed-web12.nnov.ru",
    "back::Thin_Client_Users:88,nn-sed-web13.nnov.ru",
    "back::Thin_Client_Users:88,nn-sed-web14.nnov.ru",
    "back::Thin_Client_Users:88,nn-sed-web16.nnov.ru",
    "back::Thin_Client_Users:88,BACKEND"

]


def parse_data(data):
    """
    Парсинг данных haproxy
    :param data: данные из csv haproxy в виде кучи строк
    :return: список данных парами
    """
    parsed_data = []
    row_data = []  # небольшой костыль (т.к. нужен список списка)
    lines = data.split('\n')  # сплитим на строки
    for line in lines:
        if line.startswith("front::Thin_Client:88"):  # ориентир интересующих нас данных (фронтент ТК)
            parts = line.split(',')  # сплитим по запятым на элементы
            if len(parts) >= 34:  # доп. проверка, чтобы не прочитать корявую строку
                # row_data.append(int(parts[4]))  # сохраняем результаты
                row_data.append(int(parts[-17]))
                # rate (33 элемент) - количество сессий в секунду
                # scur (4 элемент) - количество сессий текущих
                # req_rate (-17 элемент) - количество запросов в секунду
            else:
                print("Ошибка: Неправлиьный формат данных (строка не соответствует затребованным данных).")
        for server in servers_strings:  # бежим по серверам из списка servers_strings
            if line.startswith(server):
                parts = line.split(',')  # сплитим по запятым на элементы
                if len(parts) >= 34:  # доп. проверка, чтобы не прочитать корявую строку
                    row_data.append(int(parts[4]))  # сохраняем результаты
                    #
                    # scur (4 элемент) - количество сессий текущих
                    #
                else:
                    print("Ошибка: Неправлиьный формат данных (строка не соответствует затребованным данных).")
    parsed_data.append(row_data)
    return parsed_data


def write_to_excel(data):
    """
    Запись данных в таблицу excel
    :param data: список распарсенных данных
    :return:
    """
    try:
        wb = load_workbook("haproxy_stats.xlsx")  # открываем таблицу
        ws = wb.active
    except FileNotFoundError:   # если её ещё нет - создаём
        wb = Workbook()
        ws = wb.active
        # Заголовки столбцов стандартные для общего фронтента балансера
        headers = ['Time',
                   # 'balance Ses/sec',
                   # 'balancer_Ses',
                   'balancer_Req/sec']

        # Заголовки для каждого сервера
        for server in servers_strings:
            headers.append(f"{server.split(',')[-1]}_Ses")
        ws.append(headers)

    for row in data:
        row.insert(0, datetime.now().strftime("%H:%M:%S"))  # Вставка времени в начало каждой строки
        ws.append(row)

    wb.save("haproxy_stats.xlsx")


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Формат команды: python3 haproxy_collecting_stats.py <время_сбора_данных_в_минутах> <время_шага_в_секундах>")
        sys.exit(1)

    duration_in_minutes = int(sys.argv[1])  # параметр времени работы
    # duration_in_minutes = 10  # тест

    sleep_time = int(sys.argv[2])  # параметр времени задержки (шага между считываниями)

    end_time = datetime.now() + timedelta(minutes=duration_in_minutes)  # расчёт времени завершения скрипта

    # адрес csv статистики haproxy
    url = "http://10.10.130.103:81/stats;csv"

    while datetime.now() < end_time:
        data = fetch_data(url)  # получаем данные

        # data = input()  # тест

        if data:
            parsed_data = parse_data(data)  # парсинг данных
            write_to_excel(parsed_data)  # запись в таблицу
            print(f"Данные записаны ... {datetime.now().strftime('%H:%M:%S')}")
        time.sleep(sleep_time)

    print(" - - - Сбор данных завершён - - - ")

